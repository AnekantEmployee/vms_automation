import openpyxl
import pandas as pd
from components.target_columns import importantColumns
from openpyxl.styles import Font, PatternFill, Alignment

def style_excel_sheet(worksheet):
    """Apply basic styling to Excel worksheet"""
    if worksheet.max_row > 0:
        # Style header row
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                cell_length = len(str(cell.value)) if cell.value else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass

        adjusted_width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def create_observations_sheet(df):
    """
    Create observations sheet where:
    - Columns are unique values from 'Type' column
    - Each column contains data from the 'Data' column filtered by that Type
    """
    # Find Type column (case-insensitive)
    type_col = next((col for col in df.columns if col.lower() == "type"), None)
    data_col = next((col for col in df.columns if col.lower() == "data"), None)
    
    if not type_col:
        print("Warning: 'Type' column not found in data")
        print(f"Available columns: {list(df.columns)}")
        return pd.DataFrame()
    
    if not data_col:
        print("Warning: 'Data' column not found in data")
        print(f"Available columns: {list(df.columns)}")
        return pd.DataFrame()
    
    print(f"Found Type column: '{type_col}'")
    print(f"Found Data column: '{data_col}'")

    # Get unique types to create columns
    unique_types = df[type_col].dropna().unique()
    unique_types = list(set(unique_types).intersection(set(importantColumns)))

    # Create a dictionary to hold data for each type
    observations_data = {}

    for type_name in unique_types:
        # Filter rows for this type and get Data column values
        type_data_values = df[df[type_col] == type_name][data_col].dropna()
        
        # Convert to strings, strip whitespace, and remove empty strings
        all_values = [
            str(val).strip() for val in type_data_values 
            if str(val).strip()
        ]
        
        # Remove duplicates while preserving order
        unique_values = list(dict.fromkeys(all_values))
        observations_data[type_name] = unique_values if unique_values else [""]

    # Find maximum length to determine rows needed
    max_length = max(len(vals) for vals in observations_data.values()) if observations_data else 0

    # Pad lists to make them equal length
    for type_name in observations_data:
        current_length = len(observations_data[type_name])
        if current_length < max_length:
            observations_data[type_name].extend([""] * (max_length - current_length))

    # Create DataFrame
    observations_df = pd.DataFrame(observations_data)

    print(
        f"Created observations sheet with {len(observations_df)} rows and {len(observations_df.columns)} columns"
    )
    print("Values count per type:")
    for col in observations_df.columns:
        non_empty = observations_df[col].astype(str).str.strip().ne("").sum()
        print(f"- {col}: {non_empty} values")

    return observations_df