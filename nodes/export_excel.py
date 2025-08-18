import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def style_excel_sheet(worksheet):
    """Apply consistent styling to Excel sheets"""
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), 
                   right=Side(style='thin'), 
                   top=Side(style='thin'), 
                   bottom=Side(style='thin'))
    
    # Style headers
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths and apply borders
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            cell.border = border
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def export_excel_node(state):
    """
    Final node: Export all sheets to Excel file with styling
    Now includes scorecard sheet from analyze_vulnerability_node
    """
    try:
        print(f"💾 Exporting to Excel: {state['output_file']}")

        if not state.get("success", False):
            raise Exception("Previous nodes failed")

        # Get all DataFrames
        original_df = state.get("original_df", pd.DataFrame())
        observations_df = state.get("observations_df", pd.DataFrame())
        findings_df = state.get("enriched_findings_df", pd.DataFrame())
        scorecard_df = state.get("scorecard_df", pd.DataFrame())

        # Save to Excel with multiple sheets
        with pd.ExcelWriter(state["output_file"], engine="openpyxl") as writer:
            # Save original data if exists
            if not original_df.empty:
                original_df.to_excel(writer, sheet_name="Original Data", index=False)
                print(f"✓ Saved Original Data sheet ({original_df.shape[0]} rows)")

            # Save observations if exists
            if not observations_df.empty:
                observations_df.to_excel(writer, sheet_name="Observations", index=False)
                print(f"✓ Saved Observations sheet ({observations_df.shape[0]} rows)")

            # Save enriched findings
            if not findings_df.empty:
                findings_df.to_excel(writer, sheet_name="Findings", index=False)
                print(f"✓ Saved Findings sheet ({findings_df.shape[0]} rows)")

            # Save scorecard if exists
            if not scorecard_df.empty:
                scorecard_df.to_excel(writer, sheet_name="Security Scorecard", index=False)
                print(f"✓ Saved Security Scorecard sheet")

        # Apply styling to all sheets
        workbook = openpyxl.load_workbook(state["output_file"])
        
        # Special styling for scorecard
        if "Security Scorecard" in workbook.sheetnames:
            scorecard_sheet = workbook["Security Scorecard"]
            
            # Add color coding based on severity
            severity_colors = {
                "Critical": "FF0000",  # Red
                "High": "FF6600",      # Orange
                "Medium": "FFCC00",    # Yellow
                "Low": "00B0F0",       # Light blue
                "Info": "92D050"       # Green
            }
            
            for row in scorecard_sheet.iter_rows(min_row=2, max_row=6):
                severity = row[0].value
                if severity in severity_colors:
                    fill = PatternFill(start_color=severity_colors[severity],
                                     end_color=severity_colors[severity],
                                     fill_type="solid")
                    for cell in row[:4]:  # Color only first 4 columns
                        cell.fill = fill
            
            # Highlight security grade
            for row in scorecard_sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "Grade:" in cell.value:
                        grade = cell.value.split(":")[1].strip()
                        grade_color = {
                            "A": "00B050",  # Green
                            "B": "92D050",   # Light green
                            "C": "FFC000",   # Yellow
                            "D": "FF6600",   # Orange
                            "E": "FF0000",   # Red
                            "F": "C00000"    # Dark red
                        }.get(grade, "FFFFFF")
                        cell.fill = PatternFill(start_color=grade_color,
                                              end_color=grade_color,
                                              fill_type="solid")
                        cell.font = Font(bold=True, size=12)

        # Apply base styling to all sheets
        for sheet_name in workbook.sheetnames:
            style_excel_sheet(workbook[sheet_name])
            print(f"✓ Applied styling to {sheet_name} sheet")

        workbook.save(state["output_file"])
        print(f"🎉 Excel export completed: {state['output_file']}")

        return {"success": True, "error": None}

    except Exception as e:
        print(f"✗ Error in export_excel_node: {e}")
        return {"success": False, "error": str(e)}